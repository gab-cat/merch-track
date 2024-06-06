from audioop import avg
import os
from django.conf import settings
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.models import User
from app.models import Customer, CustomerSatisfactionSurvey, Fulfillment, Log, Order, OrderItem, Payment, Product
from app.forms import CustomerForm
from django.db.models import Sum, Count
from django.utils import timezone
from django.db.models.functions import TruncMonth
from django.db.models import F, Avg, Count
import openpyxl
from django.http import HttpResponse
from django.utils.timezone import make_naive
from django.contrib import messages
from django.db.models.functions import TruncDay
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator

import matplotlib
matplotlib.use('Agg')  # Use the 'Agg' backend for non-GUI rendering
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from io import BytesIO
import base64
import logging

# Set up logging
logger = logging.getLogger(__name__)

def report_index(request):
    try:
        font_path = os.path.join(settings.STATICFILES_DIRS[0], 'assets', 'montserrat.ttf')
        custom_font = fm.FontProperties(fname=font_path)

        # Update rcParams to use the custom font
        plt.rcParams['font.family'] = custom_font.get_name()
        
        # Ensure that the font is properly loaded
        fm.fontManager.addfont(font_path)
        fm._load_fontmanager(try_read_cache=False)

        # Sales data
        total_sales = Order.objects.aggregate(Sum('totalAmount'))['totalAmount__sum'] or 0
        total_orders = Order.objects.count()
        total_items_sold = OrderItem.objects.aggregate(Sum('quantity'))['quantity__sum'] or 0
        total_customers = Customer.objects.count()
        pending_orders_amount = Order.objects.filter(status='Pending').aggregate(Sum('totalAmount'))['totalAmount__sum'] or 0
        pending_orders_count = Order.objects.filter(status='Pending').count()
        collected_amount = total_sales - pending_orders_amount
        collection_rate = (collected_amount / total_sales * 100) if total_sales > 0 else 0

        # Best sellers
        best_sellers = OrderItem.objects.values('productId__name').annotate(total_sold=Sum('quantity')).order_by('-total_sold')[:3]

        # Top customers
        top_customers = Order.objects.values('customerId__user__last_name', 'customerId__user__first_name', 'customerId').annotate(total_spent=Sum('totalAmount')).order_by('-total_spent')[:3]

        # Top users
        top_users = User.objects.annotate(processed_count=Count('processed_orders')).order_by('-processed_count')[:3]

        # Customer Retention Rate
        customer_order_counts = Order.objects.values('customerId').annotate(total=Count('orderId')).order_by('-total')
        returning_customers = len([c for c in customer_order_counts if c['total'] > 1])
        retention_rate = (returning_customers / total_customers) * 100 if total_customers > 0 else 0

        # Average Order Value (AOV)
        avg_order_value = total_sales / total_orders if total_orders > 0 else 0

        # Generate charts
        sales_data = Order.objects.annotate(day=TruncDay('orderDate')).values('day').annotate(total=Sum('totalAmount')).order_by('day')
        days = [data['day'].strftime('%Y-%m-%d') for data in sales_data]
        sales = [data['total'] for data in sales_data]

        fig, ax = plt.subplots()
        ax.plot(days, sales, marker='o')
        ax.set_title('Daily Sales')
        ax.set_xlabel('Day')
        ax.set_ylabel('Sales Amount')
        plt.xticks(rotation=45)
        plt.tight_layout()

        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()
        sales_chart = base64.b64encode(image_png).decode('utf-8')

        # Monthly Revenue Growth
        monthly_revenue = Order.objects.annotate(month=TruncMonth('orderDate')).values('month').annotate(total=Sum('totalAmount')).order_by('month')
        months = [data['month'].strftime('%Y-%m') for data in monthly_revenue]
        revenue = [data['total'] for data in monthly_revenue]

        # Order Status Distribution
        order_status_distribution = Order.objects.values('status').annotate(count=Count('status'))
        statuses = [data['status'] for data in order_status_distribution]
        status_counts = [data['count'] for data in order_status_distribution]

        # Sales by Category
        sales_by_category = OrderItem.objects.values('productId__category').annotate(total=Sum('quantity')).order_by('-total')
        categories = [data['productId__category'] for data in sales_by_category]
        sales_counts = [data['total'] for data in sales_by_category]

        # Generate Monthly Revenue Growth Chart
        fig1, ax1 = plt.subplots()
        ax1.plot(months, revenue, marker='o', color='b')
        ax1.set_title('Monthly Revenue Growth')
        ax1.set_xlabel('Month')
        ax1.set_ylabel('Revenue')
        plt.xticks(rotation=45)
        plt.tight_layout()
        buffer1 = BytesIO()
        plt.savefig(buffer1, format='png')
        buffer1.seek(0)
        image_png1 = buffer1.getvalue()
        buffer1.close()
        revenue_chart = base64.b64encode(image_png1).decode('utf-8')

        # Generate Order Status Distribution Chart
        fig2, ax2 = plt.subplots()
        ax2.pie(status_counts, labels=statuses, autopct='%1.1f%%', startangle=90)
        ax2.set_title('Order Status Distribution')
        plt.tight_layout()
        buffer2 = BytesIO()
        plt.savefig(buffer2, format='png')
        buffer2.seek(0)
        image_png2 = buffer2.getvalue()
        buffer2.close()
        status_chart = base64.b64encode(image_png2).decode('utf-8')

        # Generate Sales by Category Chart
        fig3, ax3 = plt.subplots()
        ax3.bar(categories, sales_counts, color='g')
        ax3.set_title('Sales by Category')
        ax3.set_xlabel('Category')
        ax3.set_ylabel('Sales Count')
        plt.xticks(rotation=45)
        plt.tight_layout()
        buffer3 = BytesIO()
        plt.savefig(buffer3, format='png')
        buffer3.seek(0)
        image_png3 = buffer3.getvalue()
        buffer3.close()
        category_chart = base64.b64encode(image_png3).decode('utf-8')

        # Customer Satisfaction Survey Data
        survey_data = CustomerSatisfactionSurvey.objects.all()
        total_surveys = survey_data.count()
        survey_rate = (total_surveys / total_orders) * 100 if total_orders > 0 else 0

        # Average Scores
        avg_scores = survey_data.aggregate(
            avg_q1=Avg('question1'),
            avg_q2=Avg('question2'),
            avg_q3=Avg('question3'),
            avg_q4=Avg('question4')
        )

        # Distribution of Responses
        q1_distribution = survey_data.values('question1').annotate(count=Count('question1')).order_by('question1')
        q2_distribution = survey_data.values('question2').annotate(count=Count('question2')).order_by('question2')
        q3_distribution = survey_data.values('question3').annotate(count=Count('question3')).order_by('question3')
        q4_distribution = survey_data.values('question4').annotate(count=Count('question4')).order_by('question4')

        # Generate charts for each question
        def generate_chart(distribution, key, title):
            values = [item[key] for item in distribution]
            counts = [item['count'] for item in distribution]
            fig, ax = plt.subplots()
            ax.bar(values, counts)
            ax.set_title(title)
            buffer = BytesIO()
            plt.savefig(buffer, format='png')
            buffer.seek(0)
            image_png = buffer.getvalue()
            buffer.close()
            return base64.b64encode(image_png).decode('utf-8')

        q1_chart = generate_chart(q1_distribution, 'question1', 'Satisfaction with the New Order System')
        q2_chart = generate_chart(q2_distribution, 'question2', 'Ease of Use of System Features')
        q3_chart = generate_chart(q3_distribution, 'question3', 'Satisfaction with the Order Processor')
        q4_chart = generate_chart(q4_distribution, 'question4', 'Likelihood to Order Again or Recommend')

        # Fulfillment Efficiency Data
        fulfillment_efficiency = calculate_fulfillment_efficiency()

    # Exclude refunded payments and get the total amount processed by each user
        top_collectors = Payment.objects.exclude(paymentStatus='Refunded')\
                           .values('processedBy__last_name', 'processedBy__first_name', 'processedBy')\
                           .annotate(total_amount=Sum('amount'))\
                           .order_by('-total_amount')[:3]


        messages.info(request, "Welcome to Insightify v1.0.1")
        return render(request, 'reports/report_index.html', {
            'total_sales': total_sales,
            'total_orders': total_orders,
            'total_items_sold': total_items_sold,
            'total_customers': total_customers,
            'best_sellers': best_sellers,
            'top_customers': top_customers,
            'top_users': top_users,
            'collection_rate': collection_rate,
            'pending_orders_amount': pending_orders_amount,
            'pending_orders_count': pending_orders_count,
            'collected_amount': collected_amount,
            'sales_chart': sales_chart,
            'revenue_chart': revenue_chart,
            'status_chart': status_chart,
            'category_chart': category_chart,
            'retention_rate': retention_rate,
            'avg_order_value': avg_order_value,
            'returning_customers': returning_customers,
            'total_surveys': total_surveys,
            'survey_rate': survey_rate,
            'avg_scores': avg_scores,
            'q1_chart': q1_chart,
            'q2_chart': q2_chart,
            'q3_chart': q3_chart,
            'q4_chart': q4_chart,
            'fulfillment_efficiency': fulfillment_efficiency,
            'top_collectors': top_collectors, 
            'latest_logs': get_latest_logs
        })
    except Exception as e:
        logger.error(f"An error occurred: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)




def order_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        orders = Order.objects.filter(orderDate__range=[start_date, end_date])
    else:
        orders = Order.objects.all()
    
    if request.GET.get('export') == 'excel':
        return or_export_to_excel(orders)

    return render(request, 'reports/order_report.html', {'orders': orders, 'latest_logs': get_latest_logs})

def or_export_to_excel(orders):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = ['Order ID', 'Customer', 'Date', 'Status', 'Total Amount']
    ws.append(headers)

    for order in orders:
        ws.append([
            order.orderId,
            order.customerId.user.username,
            make_naive(order.orderDate),  # Convert to naive datetime
            order.status,
            order.totalAmount
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=order_report.xlsx'
    wb.save(response)
    
    return response

def product_report(request, product_id=None):
    status = request.GET.get('status')
    products = Product.objects.all()
    order_statuses = ['Pending', 'Under Production', 'Completed']

    selected_product_name=""
    
    if product_id:
        orders = Order.objects.filter(orderitem__productId=product_id)
        if status:
            orders = orders.filter(status=status)
        orders = orders.distinct()
        total_quantity = OrderItem.objects.filter(productId=product_id).aggregate(Sum('quantity'))['quantity__sum'] or 0

                # Get the selected product's name
        selected_product = Product.objects.filter(productId=product_id).first()
        if selected_product:
            selected_product_name = selected_product.name

    else:
        orders = Order.objects.none()
        total_quantity = 0

    if request.GET.get('export') == 'excel':
        return pr_export_to_excel(orders)

    return render(request, 'reports/product_report.html', {
        'orders': orders,
        'total_quantity': total_quantity,
        'products': products,
        'selected_product_id': product_id,
        'selected_product_name': selected_product_name,
        'status': status,
        'order_statuses': order_statuses,
        'latest_logs': get_latest_logs
    })

def pr_export_to_excel(orders):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = ['Order ID', 'Customer', 'Date', 'Email', 'Phone', 'Quantity', 'Status', 'Size', 'Note']
    ws.append(headers)

    for order in orders:
        for item in order.orderitem_set.all():
            ws.append([
                order.orderId,
                f"{order.customerId.user.first_name} {order.customerId.user.first_name}",
                make_naive(order.orderDate),  # Convert to naive datetime
                order.customerId.user.email,
                order.customerId.phone,
                item.quantity,
                order.status,
                item.size,
                item.customerNote
            ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=product_report.xlsx'
    wb.save(response)
    
    return response

def sales_data_api(request, product_id):
    sales_over_time = (
        OrderItem.objects.filter(productId=product_id)
        .values('orderId__orderDate')
        .annotate(total_quantity=Sum('quantity'))
        .order_by('orderId__orderDate')
    )
    quantity_by_customer = (
        OrderItem.objects.filter(productId=product_id)
        .values('orderId__customerId__user__username')
        .annotate(total_quantity=Sum('quantity'))
        .order_by('-total_quantity')
    )
    data = {
        'sales_over_time': list(sales_over_time),
        'quantity_by_customer': list(quantity_by_customer),
        'latest_logs': get_latest_logs
    }
    return JsonResponse(data)

def customer_report(request):
    customer_id = request.GET.get('customer_id')
    status = request.GET.get('status')
    customers = Customer.objects.all()
    order_statuses = ['Pending', 'Under Production', 'Completed']
    
    if customer_id:
        orders = Order.objects.filter(customerId=customer_id)
        if status:
            orders = orders.filter(status=status)
        orders = orders.distinct()
        total_spending = orders.aggregate(Sum('totalAmount'))['totalAmount__sum'] or 0
    else:
        orders = Order.objects.none()
        total_spending = 0

    if request.GET.get('export') == 'excel':
        return export_to_excel(orders)

    return render(request, 'reports/customer_report.html', {
        'orders': orders,
        'total_spending': total_spending,
        'customers': customers,
        'selected_customer_id': customer_id,
        'status': status,
        'order_statuses': order_statuses,
    })

def export_to_excel(orders):
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = ['Order ID', 'Customer', 'Date', 'Status', 'Total Amount']
    ws.append(headers)

    for order in orders:
        ws.append([
            order.orderId,
            order.customerId.user.username,
            make_naive(order.orderDate),  # Convert to naive datetime
            order.status,
            order.totalAmount
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_report.xlsx'
    wb.save(response)
    
    return response

def sales_report(request):
    start_date = request.GET.get('start_date', timezone.now().replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))

    orders = Order.objects.filter(orderDate__range=[start_date, end_date])
    total_sales = orders.aggregate(total_sales=Sum('totalAmount'))['total_sales'] or 0
    total_discount = orders.aggregate(total_discount=Sum('discountAmount'))['total_discount'] or 0
    total_orders = orders.count()
    avg_order_value = total_sales / total_orders if total_orders > 0 else 0

    if request.GET.get('export') == 'excel':
        return sr_export_to_excel(orders, total_sales, total_discount, total_orders, avg_order_value, start_date, end_date)

    return render(request, 'reports/sales_report.html', {
        'orders': orders,
        'total_sales': total_sales,
        'total_discount': total_discount,
        'total_orders': total_orders,
        'avg_order_value': avg_order_value,
        'start_date': start_date,
        'end_date': end_date,
        'latest_logs': get_latest_logs
    })

def sr_export_to_excel(orders, total_sales, total_discount, total_orders, avg_order_value, start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    headers = ['Order ID', 'Customer', 'Date', 'Status', 'Total Amount', 'Discount Amount']
    ws.append(headers)

    for order in orders:
        ws.append([
            order.orderId,
            order.customerId.user.username,
            make_naive(order.orderDate),  # Convert to naive datetime
            order.status,
            order.totalAmount,
            order.discountAmount
        ])

    # Summary row
    ws.append([])
    ws.append(['Total Sales', total_sales])
    ws.append(['Total Discount', total_discount])
    ws.append(['Total Orders', total_orders])
    ws.append(['Average Order Value', avg_order_value])
    ws.append(['Start Date', start_date])
    ws.append(['End Date', end_date])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
    wb.save(response)
    
    return response

def fulfillment_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    fulfillments = Fulfillment.objects.all()
    if start_date and end_date:
        start_date = parse_date(start_date)
        end_date = parse_date(end_date)
        fulfillments = fulfillments.filter(fulfillmentDate__range=[start_date, end_date])
    
    if request.GET.get('export') == 'excel':
        return fulfillment_export_to_excel(fulfillments)

    return render(request, 'reports/fulfillment_report.html', {'fulfillments': fulfillments, 'start_date': start_date, 'end_date': end_date, 'latest_logs': get_latest_logs})

def fulfillment_export_to_excel(fulfillments):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fulfillment Report"

    headers = ['Fulfillment ID', 'Order ID', 'Customer', 'Processed By', 'Fulfillment Date', 'Estimated Delivery Date', 'Status']
    ws.append(headers)

    for fulfillment in fulfillments:
        ws.append([
            fulfillment.fulfillmentId,
            fulfillment.orderId.orderId,
            f"{fulfillment.orderId.customerId.user.first_name} {fulfillment.orderId.customerId.user.last_name}",
            f"{fulfillment.processedBy.first_name} {fulfillment.processedBy.last_name}",
            make_naive(fulfillment.fulfillmentDate),
            make_naive(fulfillment.orderId.estimatedDeliveryDate),
            fulfillment.status
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=fulfillment_report.xlsx'
    wb.save(response)
    return response

def calculate_fulfillment_efficiency():
    # Calculate on-time and late fulfillments based on status
    on_time_fulfillments = Fulfillment.objects.filter(status='On Time').count()
    late_fulfillments = Fulfillment.objects.filter(status='Late').count()
    total_fulfillments = Fulfillment.objects.count()
    
    # Calculate efficiency
    efficiency = (on_time_fulfillments / total_fulfillments) * 100 if total_fulfillments > 0 else 0
    
    # Calculate unfulfilled orders
    unfulfilled_orders = Order.objects.filter(status='Ready').count()

    return {
        'on_time_fulfillments': on_time_fulfillments,
        'late_fulfillments': late_fulfillments,
        'efficiency': efficiency,
        'total_fulfillments': total_fulfillments,
        'unfulfilled_orders': unfulfilled_orders,
        'total_fulfillments': total_fulfillments,
    }

def survey_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    surveys = CustomerSatisfactionSurvey.objects.all()
    if start_date and end_date:
        start_date = parse_date(start_date)
        end_date = parse_date(end_date)
        surveys = surveys.filter(submit_date__range=[start_date, end_date])
    
    if request.GET.get('export') == 'excel':
        return sr_export_to_excel(surveys, start_date, end_date)

    return render(request, 'reports/survey_report.html', {'surveys': surveys, 'start_date': start_date, 'end_date': end_date, 'latest_logs': get_latest_logs})

def sr_export_to_excel(surveys, start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Survey Report"

    headers = ['Order ID', 'Submit Date', 'Q1: Satisfaction', 'Q2: Ease of Use', 'Q3: Processor Satisfaction', 'Q4: Likelihood to Recommend', 'Comments']
    ws.append(headers)

    for survey in surveys:
        ws.append([
            survey.order.orderId,
            make_naive(survey.submit_date),
            survey.question1,
            survey.question2,
            survey.question3,
            survey.question4,
            survey.comments
        ])

    # Summary row
    avg_scores = surveys.aggregate(
        avg_q1=Avg('question1'),
        avg_q2=Avg('question2'),
        avg_q3=Avg('question3'),
        avg_q4=Avg('question4')
    )
    
    ws.append([])
    ws.append(['Average Scores', '', avg_scores['avg_q1'], avg_scores['avg_q2'], avg_scores['avg_q3'], avg_scores['avg_q4'], ''])
    
    # Date Range
    if start_date and end_date:
        ws.append([])
        ws.append(['Start Date', start_date.strftime('%Y-%m-%d')])
        ws.append(['End Date', end_date.strftime('%Y-%m-%d')])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=survey_report.xlsx'
    wb.save(response)
    
    return response


def collections_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    payments = Payment.objects.all()
    if start_date and end_date:
        start_date = parse_date(start_date)
        end_date = parse_date(end_date)
        payments = payments.filter(paymentDate__range=[start_date, end_date])
    
    if request.GET.get('export') == 'excel':
        return collections_export_to_excel(payments)

    # Exclude refunded payments and get the total amount processed by each user
    top_users = Payment.objects.exclude(paymentStatus='Refunded')\
                           .values('processedBy__last_name', 'processedBy__first_name', 'processedBy')\
                           .annotate(total_amount=Sum('amount'))\
                           .order_by('-total_amount')[:3]

    # Total collection by payment method excluding refunded payments
    total_collection_by_method = Payment.objects.exclude(paymentStatus='Refunded')\
                                            .values('paymentMethod')\
                                            .annotate(total_amount=Sum('amount'))

    return render(request, 'reports/collections_report.html', {
        'payments': payments,
        'start_date': start_date,
        'end_date': end_date,
        'top_users': top_users,
        'total_collection_by_method': total_collection_by_method,
        'latest_logs': get_latest_logs
    })

def collections_export_to_excel(payments):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collections Report"

    headers = ['Payment ID', 'Order ID', 'Customer', 'Processed By', 'Payment Date', 'Amount', 'Payment Method', 'Payment Status', 'Reference Number']
    ws.append(headers)

    for payment in payments:
        ws.append([
            payment.paymentId,
            payment.orderId.orderId,
            f"{payment.customerId.user.first_name} {payment.customerId.user.last_name}",
            f"{payment.processedBy.first_name} {payment.processedBy.last_name}",
            make_naive(payment.paymentDate),
            payment.amount,
            payment.paymentMethod,
            payment.paymentStatus,
            payment.referenceNumber
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=collections_report.xlsx'
    wb.save(response)
    
    return response


def log_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    created_by = request.GET.get('created_by')
    logs = Log.objects.all()

    if start_date and end_date:
        start_date = parse_date(start_date)
        end_date = parse_date(end_date)
        logs = logs.filter(created_date__range=[start_date, end_date])
    
    if created_by:
        logs = logs.filter(created_by__username__icontains=created_by)

    # Pagination
    paginator = Paginator(logs, 10)  # Show 10 logs per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.GET.get('export') == 'excel':
        return log_export_to_excel(logs)

    return render(request, 'reports/log_report.html', {
        'logs': logs,
        'page_obj': page_obj,
        'start_date': start_date,
        'end_date': end_date,
        'created_by': created_by,
        'latest_logs': get_latest_logs,
    })

def log_export_to_excel(logs):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Log Report"

    headers = ['Log ID', 'Created Date', 'Created By', 'Reason', 'System Text']
    ws.append(headers)

    for log in logs:
        ws.append([
            log.logId,
            log.created_date.strftime('%Y-%m-%d %H:%M:%S'),
            f"{log.created_by.first_name} {log.created_by.last_name}",
            log.reason,
            log.system_text
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=log_report.xlsx'
    wb.save(response)


    return response


def get_latest_logs(count=5):
    return Log.objects.order_by('-created_date')[:count]
